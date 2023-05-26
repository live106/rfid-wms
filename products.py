from PyQt5.QtWidgets import QTableWidget, QLabel, QStackedWidget, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFileDialog, QMessageBox, QSpinBox
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5 import QtGui
from openpyxl import load_workbook, Workbook # for excel import and export
import database # for data operations

class PageLink(QLabel):
    # Signal emitted when label is clicked
    clicked = pyqtSignal(str)

    def __init__(self, text, parent=None):
        super().__init__(text, parent=parent)
        self.setTextInteractionFlags(Qt.LinksAccessibleByMouse)
        self.setStyleSheet("color: blue;") # set text color to blue to emulate a link
        self.setCursor(Qt.PointingHandCursor) # set the cursor to link pointer

    def mousePressEvent(self, event):
        self.clicked.emit(self.text()) # emit the clicked signal when pressed
        return super().mousePressEvent(event)

class Products(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.layout = QVBoxLayout(self)

        # create a button row with excel import, excel template download and delete buttons
        self.button_row = QHBoxLayout()
        self.import_button = QPushButton("Import Products")
        self.import_button.setFont(QtGui.QFont("Arial", 16))
        self.import_button.clicked.connect(self.import_excel)
        self.template_button = QPushButton("Download Template")
        self.template_button.setFont(QtGui.QFont("Arial", 16))
        self.template_button.clicked.connect(self.download_template)
        self.delete_button = QPushButton("Delete")
        self.delete_button.setFont(QtGui.QFont("Arial", 16))
        self.delete_button.clicked.connect(self.delete_product)
        
        self.button_row.addWidget(self.import_button)
        self.button_row.addWidget(self.template_button)
        self.button_row.addWidget(self.delete_button)
        self.layout.addLayout(self.button_row)

        # create the stacked widget that will contain each page of table
        self.stack_widget = QStackedWidget(parent=self)
        self.layout.addWidget(self.stack_widget)

        # setup the layout for the page numbers below the stacked widget
        self.pagination_layout = QHBoxLayout()
        self.pagination_layout.addStretch(0)
        self.prev_label = QLabel("<")
        self.pagination_layout.addWidget(self.prev_label)

        # create pages and corresponding labels
        self.page_links = []
        
        # set the products per page to 10 by default
        self.products_per_page = 20
        
        # create a spin box to change the products per page
        self.spin_box = QSpinBox()
        self.spin_box.setMinimum(1) # minimum value is 1
        self.spin_box.setMaximum(50) # maximum value is 50
        self.spin_box.setValue(self.products_per_page) # initial value
        self.spin_box.valueChanged.connect(self.change_products_per_page) # connect the value changed signal to a slot
        
        # add the spin box to the pagination layout
        self.pagination_layout.addWidget(QLabel("Per Page: "))
        self.pagination_layout.addWidget(self.spin_box)

        self.products = database.get_products() # get all products from database
        
        self.next_label = None
        # create pages according to the products per page value
        self.create_pages()

    def create_pages(self):
        self.clear_pages() # clear all pages before creating new ones
        # calculate how many pages are needed
        self.total_pages = (len(self.products) - 1) // self.products_per_page + 1 
        
        for i in range(1, self.total_pages + 1):
            page_link = PageLink(str(i), parent=self)
            self.pagination_layout.addWidget(page_link)
            page = QTableWidget()
            page.setColumnCount(2) # barcode and name columns
            page.setHorizontalHeaderLabels(["Barcode", "Product Name"])
            page.setRowCount(self.products_per_page) # set the row count to products per page
            page.verticalHeader().hide() # hide the vertical header
            page.horizontalHeader().setStretchLastSection(True) # stretch the last column to fill the width
            page.setEditTriggers(QTableWidget.NoEditTriggers) # disable editing
            page.setSelectionBehavior(QTableWidget.SelectRows) # enable row selection
            page.setSelectionMode(QTableWidget.SingleSelection) # only allow single selection

            # fill the table with products data according to the page number
            start_index = (i - 1) * self.products_per_page # the start index of the products list for this page
            end_index = min(start_index + self.products_per_page, len(self.products)) # the end index of the products list for this page
            for j in range(start_index, end_index):
                barcode, name = self.products[j] # get the product data from the list
                barcode_item = QLabel(barcode) # create a label for barcode
                barcode_item.setAlignment(Qt.AlignCenter) # align the label to center
                name_item = QLabel(name) # create a label for name
                name_item.setAlignment(Qt.AlignCenter) # align the label to center
                page.setCellWidget(j - start_index, 0, barcode_item) # set the barcode label to the first column
                page.setCellWidget(j - start_index, 1, name_item) # set the name label to the second column
            
            self.stack_widget.addWidget(page) # add the page to the stacked widget
            page_link.clicked.connect(self.switch_page) # connect the page link signal to the switch page slot
        
        if not self.next_label:
            self.next_label = QLabel(">")
            self.pagination_layout.addWidget(self.next_label)
            self.layout.addLayout(self.pagination_layout)

    def switch_page(self, page):
        # switch to the corresponding page index when a page link is clicked
        self.stack_widget.setCurrentIndex(int(page) - 1)

    def import_excel(self):
        # import data from an excel file and insert them into database and table widget
        file_name = QFileDialog.getOpenFileName(self, "Choose Excel file", "", "Excel file (*.xlsx)") # get the file name from a file dialog
        if file_name[0]:
            workbook = load_workbook(file_name[0]) # open the workbook
            sheet = workbook.active # get the active sheet
            for row in sheet.iter_rows(min_row=2): # iterate over the rows, skipping the first row (headers)
                barcode = row[0].value # get the barcode value from the first column
                name = row[1].value # get the name value from the second column
                database.add_product(barcode, name) # add the product to database
                print(barcode, name)

            QMessageBox.information(self, "SCUCCESS", "Import successful !") # show a message box to indicate success

            # refresh the table widget with new data
            self.products = database.get_products() 
            self.create_pages() # recreate pages according to new data

    def download_template(self):
        # export a template excel file with headers
        file_name = QFileDialog.getSaveFileName(self, "Save Excel file", "", "Excel file (*.xlsx)") # get the file name from a save file dialog
        if file_name[0]:
            workbook = Workbook() # create a workbook
            sheet = workbook.active # get the active sheet
            headers = ["Barcode", "Product Name"] # set the headers list
            for i in range(len(headers)): # iterate over the headers list
                sheet.cell(row=1, column=i+1).value = headers[i] # write the header to the first row of the sheet
            
            workbook.save(file_name[0]) # save the workbook to the file name

    def delete_product(self):
        # delete the selected product from database and table widget
        
        current_page = self.stack_widget.currentIndex() + 1 # get the current page number
        current_row = self.stack_widget.currentWidget().currentRow() # get the current row number of the table widget
        
        if current_row == -1: 
          QMessageBox.warning(self, "WARNING", "Please select the product record !") 
          return
        
        index = (current_page - 1) * self.products_per_page + current_row 
        barcode, name = self.products[index] 
        
        reply = QMessageBox.question(self, "Confirm", f"Delete {barcode} {name} ?", QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
          database.delete_product(barcode) 
          QMessageBox.information(self, "SUCCESS", f"Delete {barcode} {name} successful !") 
          self.products = database.get_products() 
          self.create_pages()

    def change_products_per_page(self, value):
        # change the products per page value and recreate pages
        self.products_per_page = value
        self.clear_pages()
        self.create_pages()

    def clear_pages(self):
        # clear all pages from the stacked widget
        while self.stack_widget.count() > 0: # loop until there is no page left
            widget = self.stack_widget.widget(0) # get the first page
            self.stack_widget.removeWidget(widget) # remove it from the stacked widget
            widget.deleteLater() # delete it later

        for i in reversed(range(self.pagination_layout.count())):
            widget = self.pagination_layout.itemAt(i).widget()
            if isinstance(widget, PageLink):
                widget.deleteLater()            

        # clear the previous pages and page links
        self.page_links.clear()

