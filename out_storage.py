from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton, QMessageBox, QLineEdit, QComboBox, QHBoxLayout, QTableWidget, QFrame, QFileDialog, QTableWidgetItem, QLCDNumber, QHeaderView
from PyQt5.QtCore import pyqtSlot, QThread, pyqtSignal, QObject, Qt
from PyQt5 import QtGui
import openpyxl
from openpyxl.styles import Alignment, numbers, Font
from openpyxl.utils import get_column_letter
import xlwings as xw
import xlsxwriter
import threading
import time
import datetime
from html import unescape

from auto_express import create_express_printer, Express
from database import *
import rfid_api
from rfid_api import stop_async_inventory_event
from config import ORDER_FOR_EXPRESS_PATH
from string_utils import halfwidth_to_fullwidth
from loading_screen import LoadingScreen

class PrintThread(QThread):
    finished = pyqtSignal(bool)

    def __init__(self, printer, orders):
        super().__init__()
        self.printer = printer
        self.orders = orders
        self._stop = False

    def run(self):
        result = self.printer.print_express_with_form(self.orders)
        self.finished.emit(result)

    def stop(self):
        self._stop = True
        
class OrderWorker(QObject):
    finished = pyqtSignal()
    result = pyqtSignal(list)

    def __init__(self, order_no_or_picking_no):
        super().__init__()
        self.order_no_or_picking_no = order_no_or_picking_no

    def run(self):
        result = get_orders_by_order_no(self.order_no_or_picking_no)
        if not result:
            result = get_orders_by_picking_no(self.order_no_or_picking_no)
        self.result.emit(result)
        self.finished.emit()

class OutboundWorker(QObject):
    finished = pyqtSignal()
    result = pyqtSignal(list)

    def __init__(self, picking_nos):
        super().__init__()
        self.picking_nos = picking_nos

    def run(self):
        result = get_orders_by_picking_nos(self.picking_nos)
        self.result.emit(result)
        self.finished.emit()

class OutStorage(QWidget):

    counter_updated = pyqtSignal(int)

    def __init__(self):
        super().__init__()
        self.order_thread = None
        self.order_worker = None

        self.outbound_thread = None
        self.outbound_worker = None

        self.inventoring = False 
        self.epc_list = []
        self.inventory_duration_ref = [0]
        self.counter_updated.connect(self.updateCounter)

        # Create a thread to update the counter every N seconds
        self.counter_thread = threading.Thread(target=self.updateCounterThread)
        # self.counter_thread.daemon = True

        # Create a thread to call rfid_api.get_epc_list
        self.rfid_thread = threading.Thread(target=self.getEpcListThread)
        # self.rfid_thread.daemon = True

        self.current_order_match_data = []
        self.all_match = False

        self.default_shipper_data = get_shipper_data()

        self.initUI()

    def initUI(self):
        '''
        # Create a subpage for 出库管理
        # self.label = QLabel("出库管理")
        # self.label.setFixedHeight(30)
        # self.label.setFont(QtGui.QFont("Arial", 20, QtGui.QFont.Bold))

        self.print_button = QPushButton("自动打印快递单功能测试", self)
        self.print_button.setFont(QtGui.QFont("Arial", 20))
        self.print_button.clicked.connect(self.print_express)

        vbox = QVBoxLayout()
        # vbox.addWidget(self.label)
        vbox.addWidget(self.print_button)
        self.setLayout(vbox)
        '''

        button_height = 64

        # Add the top row of buttons and input components
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText('PickingNo/OrNO')
        self.order_input.setFont(QtGui.QFont("Arial", button_height))
        self.order_input.setFocus(True)
        self.order_input.textChanged.connect(self.on_order_input_changed)

        # Create a button to start inventory
        self.inventory_button = QPushButton("START")
        self.inventory_button.setFont(QtGui.QFont("Arial", button_height))
        self.inventory_button.clicked.connect(self.startAsyncInventory)

        # Create a text counter to show the number of characters in the product name input box
        self.counter = QLCDNumber()

        self.express_combo = QComboBox()
        self.express_combo.setPlaceholderText('Express')
        self.express_combo.setFont(QtGui.QFont("Arial", button_height))

        self.outbound_button = QPushButton("OUTBOUND")
        self.outbound_button.setFont(QtGui.QFont("Arial", button_height))
        # self.outbound_button.setEnabled(False)
        self.outbound_button.clicked.connect(self.outbound)

        top_hbox = QHBoxLayout()
        top_hbox.addWidget(self.order_input, 4)
        top_hbox.addWidget(self.inventory_button, 4)
        top_hbox.addWidget(self.counter, 2)

        # Add the first table
        self.match_table = QTableWidget()
        self.match_table.setColumnCount(5)
        self.match_table.setHorizontalHeaderLabels(["PickingNo", "JAN", "Qty", "CurrentQty", "Matching"])
        self.match_table.setFixedHeight(260)
        self.match_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # Add the separator and second set of buttons and table
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)

        import_button = QPushButton("Import Orders")
        import_button.setFont(QtGui.QFont("Arial", 14))
        import_button.clicked.connect(self.import_data)

        template_button = QPushButton("Download Template")
        template_button.setFont(QtGui.QFont("Arial", 14))
        template_button.clicked.connect(self.download_template)

        self.order_table = QTableWidget()
        # headers = ['Type', 'JAN', 'Expiration', 'Qty', 'ZIP', 'Address', 'Name', 'TEL', 'Comment1', 'Comment2', 'Comment3', 'Comment4', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'PickingNo', 'OrNO', 'OutboundStatus', 'Express', 'ExpressNo', 'ExpressTime']
        headers = ['OutboundStatus', 'PickingNo', 'OrNO', 'Type', 'JAN', 'Expiration', 'Qty', 'ZIP', 'Address', 'Name', 'TEL', 'Comment1', 'Comment2', 'Comment3', 'Comment4', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'Express', 'ExpressNo', 'ExpressTime']
        self.order_table.setColumnCount(len(headers))
        self.order_table.setHorizontalHeaderLabels(headers)
        # self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.order_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)        

        vbox = QVBoxLayout()
        # vbox.addWidget(self.label)
        vbox.addLayout(top_hbox)

        top_hbox = QHBoxLayout()
        top_hbox.addWidget(self.express_combo, 4)
        top_hbox.addWidget(self.outbound_button, 6)
        vbox.addLayout(top_hbox)
        
        vbox.addWidget(self.match_table)
        vbox.addWidget(separator)

        top_hbox = QHBoxLayout()
        top_hbox.addWidget(import_button, 1)
        top_hbox.addWidget(template_button, 1)

        vbox.addLayout(top_hbox)
        vbox.addWidget(self.order_table)
        self.setLayout(vbox)

        self.update_express_options()
        self.reload_orders()

        self.loading_screen = LoadingScreen()

    def update_express_options(self):
        configs = get_all_express_configs()
        self.express_combo.clear()
        for config in configs:
            self.express_combo.addItem(config['alias'])

    def showEvent(self, event):
        # Set focus to order_input when the UI is shown
        # self.order_input.setEnabled(True)
        self.order_input.setFocus()
        self.update_express_options()
        self.default_shipper_data = get_shipper_data()
        event.accept()   

    @pyqtSlot()
    def print_express(self, orders):
        printer = create_express_printer(self.express_combo.currentText())
        self.order_thread = PrintThread(printer, orders)
        self.order_thread.finished.connect(self.on_print_finished)
        self.order_thread.printer.update_loading_text.connect(self.update_loading_text)
        self.order_thread.start()

    def update_loading_text(self, text):
        self.loading_screen.update_loading_text(text)
        self.setEnabled(not text)
        
    def on_print_finished(self, result):
        picking_no = self.current_order_match_data[0].get('PickingNo', None)
        if result:
            if len(self.current_order_match_data) > 0:
                if picking_no:
                    current_date = datetime.date.today()
                    formatted_date = current_date.strftime("%Y/%m/%d")
                    update_order(picking_no, {'OutboundStatus': 'Done', 'Express': self.express_combo.currentText(), 'ExpressTime': formatted_date})
                    update_epcs(self.epc_list, {'picking_no': picking_no})
                    self.reload_orders()

            self.order_input.clear()
            self.counter.display(0)
            QMessageBox.information(self, "Success", f"Order: {picking_no} Outbound Successful !")
        else:
            QMessageBox.warning(self, "Warning", f"Order: {picking_no} Outbound Failure, Please Retry !")
        # self.order_input.setEnabled(True)
        self.order_input.setFocus()

    def closeEvent(self, event):
        if self.order_thread is not None:
            self.order_thread.stop()
            self.order_thread.wait()
        event.accept()            
        
    def download_template(self):
        # Prompt the user to select a location to save the file
        file_path, _ = QFileDialog.getSaveFileName(None, "Save Template", "OrderTemplate.xlsx", "Excel Files (*.xlsx)")

        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add the desired headers to the worksheet
        # headers = ['Type', 'JAN', 'Expiration', 'ZIP', 'Address', 'Name', 'TEL', 'Text1', 'Text2', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'CustomerOrderID', 'Qty', 'PickingNo', 'Express']
        headers = ['Type', 'JAN', 'Expiration', 'Qty', 'ZIP', 'Address', 'Name', 'TEL', 'Comment1', 'Comment2', 'Comment3', 'Comment4', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'PickingNo', 'OrNO']

        for i, header in enumerate(headers):
            cell = worksheet.cell(row=1, column=i+1)
            cell.value = header

        # Save the workbook to the selected location
        workbook.save(file_path)

    def import_data(self):
        # Prompt the user to select an Excel file to import
        file_path, _ = QFileDialog.getOpenFileName(None, "Import Data", "", "Excel Files (*.xlsx)")

        # Load the workbook and select the first worksheet
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # Get the headers from the worksheet
        headers = [cell.value for cell in worksheet[1]]

        # Compare the headers to the headers of the corresponding table in your UI
        # expected_headers = ['Type', 'JAN', 'Expiration', 'ZIP', 'Address', 'Name', 'TEL', 'Text1', 'Text2', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'CustomerOrderID', 'Qty', 'PickingNo', 'Express']
        expected_headers = ['Type', 'JAN', 'Expiration', 'Qty', 'ZIP', 'Address', 'Name', 'TEL', 'Comment1', 'Comment2', 'Comment3', 'Comment4', 'D_Date', 'D_Time', 'ShipperZIP', 'ShipperName', 'ShipperAddress', 'ShipperTel', 'PickingNo', 'OrNO']

        if headers != expected_headers:
            QMessageBox.warning(None, "Error", "The headers in the selected file do not match the expected headers.")
            return
        
        # 需要转换为全角的列
        full_width_headers = ['Address', 'Name', 'Comment1', 'Comment2', 'Comment3', 'Comment4', 'ShipperName', 'ShipperAddress']
        required_headers = ['JAN', 'Qty', 'Address', 'TEL', 'PickingNo', 'OrNO']

        # Import the data from the worksheet into the database
        data = []
        for row in worksheet.iter_rows(min_row=2):
            item = {}
            for index, cell in enumerate(row):
                if expected_headers[index] in required_headers and not cell.value:
                    QMessageBox.warning(None, "Error", f'{expected_headers[index]} can not be empty!')
                    return
                if expected_headers[index] in full_width_headers:
                    item[headers[index]] = halfwidth_to_fullwidth(str(cell.value)) if cell.value else ''
                else:
                    item[headers[index]] = str(cell.value) if cell.value else ''
                

            data.append(item)
        add_orders(data)

        # Reload the orders in the UI
        self.reload_orders()

    def reload_orders(self):
        orders = get_orders()
        self.order_table.setRowCount(len(orders))
        for i, order in enumerate(orders):
            for j in range(self.order_table.columnCount()):
                header = self.order_table.horizontalHeaderItem(j)
                key = header.text()
                value = order.get(key, '')
                item = QTableWidgetItem(str(value))
                self.order_table.setItem(i, j, item)

                if key == 'OutboundStatus' and value == 'Done':
                    item.setBackground(QtGui.QColor(0, 255, 0))

    @pyqtSlot()
    def on_order_input_changed(self):
        # self.order_input.setEnabled(False)
        self.current_order_match_data.clear()
        order_no = self.order_input.text()
        if order_no:
            self.order_thread = QThread()
            self.order_worker = OrderWorker(order_no)
            self.order_worker.moveToThread(self.order_thread)
            self.order_thread.started.connect(self.order_worker.run)
            self.order_worker.finished.connect(self.order_thread.quit)
            self.order_worker.finished.connect(self.order_worker.deleteLater)
            self.order_thread.finished.connect(self.order_thread.deleteLater)
            self.order_worker.result.connect(self.update_match_table)
            self.order_thread.start()
        else:
            self.match_table.setRowCount(0)

    def update_match_table(self, result):
        self.current_order_match_data = result
        self.match_table.setRowCount(len(self.current_order_match_data))
        self.all_match = True
        for i, row in enumerate(result):
            self.match_table.setItem(i, 0, QTableWidgetItem(row["PickingNo"]))
            self.match_table.setItem(i, 1, QTableWidgetItem(row["JAN"]))
            self.match_table.setItem(i, 2, QTableWidgetItem(str(row["Qty"])))
            self.match_table.setItem(i, 3, QTableWidgetItem(str(row.get('CurrentQty', 0))))
            self.match_table.setItem(i, 4, QTableWidgetItem(str(row.get('Matching', False))))
            match = row["Qty"] == row.get('CurrentQty', 0)
            self.match_table.setItem(i, 4, QTableWidgetItem(str(match)))
            if not match:
                self.all_match = False
            if i == 0:
                alias = row.get('Type', None)
                if alias:
                    self.express_combo.setCurrentText(alias)
                else:
                    config = get_express_config_by_name(Express.SAGAWAEXP.value)
                    self.express_combo.setCurrentText(config.get('alias', ''))

        # self.outbound_button.setEnabled(self.all_match)
        # self.order_input.setEnabled(True)
        self.order_input.setFocus()

    def startAsyncInventory(self):
        self.counter_updated.emit(0)
        stop_async_inventory_event.clear()  # Unset the threading.Event object to resume the inventory process
        # self.inventory_button.setEnabled(False)
        self.inventory_button.disconnect()
        self.inventory_button.setText(f"STOP")  # Add this line to set the text of the inventory_button to "结束盘点"
        self.inventory_button.clicked.connect(self.stopAsyncInventory)  # Add this line to set the stop_async_inventory_event value to True when the button is clicked        
        if not self.counter_thread.is_alive():
            self.counter_thread = threading.Thread(target=self.updateCounterThread)
        if not self.rfid_thread.is_alive():
            self.rfid_thread = threading.Thread(target=self.getEpcListThread)
        self.inventoring = True
        self.counter_thread.start()
        self.rfid_thread.start()  # start the rfid_thread        

    def stopAsyncInventory(self):
        stop_async_inventory_event.set()  # Set the threading.Event object to True to stop the inventory process
        self.inventoring = False
        self.inventory_duration_ref = [0]
        self.counter_thread.join()
        
        self.inventory_button.setText("START")
        self.inventory_button.clicked.connect(self.startAsyncInventory)

    def getEpcListThread(self):
        rfid_api.async_get_epc_list(self.epc_list, self.inventory_duration_ref)  # call rfid_api.async_get_epc_list in a new thread
        self.counter_updated.emit(len(self.epc_list))
        self.stopAsyncInventory()       

    def updateCounterThread(self):
        while self.inventoring:
            try:
                time.sleep(1)
                self.counter_updated.emit(len(self.epc_list))
                self.inventory_button.setText(f"STOP[{(str(round(self.inventory_duration_ref[0])) + 's') if self.inventory_duration_ref[0] > 0 else ''}]")  # Add this line to set the text of the inventory_button to "结束盘点"
                print(f"Counter updated: {len(self.epc_list)}, inventory_duration_ref: {self.inventory_duration_ref[0]}")
                if not self.counter_thread.is_alive():
                    print("Counter thread terminated.")
                    break
            except Exception as e:
                error_message = str(e)
                print(error_message)
                pass
        print("Counter updated done.")

    def updateCounter(self, value):
        self.counter.display(value)
        if len(self.epc_list) > 0:
            result = get_valid_epc_barcode_counts(self.epc_list)
            print('result: ', result)
            for v in self.current_order_match_data:
                v['CurrentQty'] = result.get(v['JAN'], 0)
            self.update_match_table(self.current_order_match_data)

    def outbound(self):
        if not self.order_input.text():
            return
        if not self.express_combo.currentText():
            QMessageBox.warning(self, "Warning", "Express not selected !")
            return
        if not self.all_match:
            reply = QMessageBox.question(self, "Confirm", f"Quantity not match, confirm outbound ?", QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
        picking_nos = []
        for row in range(self.match_table.rowCount()):
            order_number_item = self.match_table.item(row, 0)
            if order_number_item is not None:
                picking_nos.append(order_number_item.text())

        if picking_nos:
            self.outbound_thread = QThread()
            self.outbound_worker = OutboundWorker(picking_nos)
            self.outbound_worker.moveToThread(self.outbound_thread)
            self.outbound_thread.started.connect(self.outbound_worker.run)
            self.outbound_worker.finished.connect(self.outbound_thread.quit)
            self.outbound_worker.finished.connect(self.outbound_worker.deleteLater)
            self.outbound_thread.finished.connect(self.outbound_thread.deleteLater)
            self.outbound_worker.result.connect(self.handle_outbound_result)            
            self.outbound_thread.start()

    def handle_outbound_result(self, outbound_orders):
        for order in outbound_orders:
            if order['OutboundStatus'] == 'Done':
                 reply = QMessageBox.question(self, "Confirm", f"Order {order['PickingNo']} outbounded already, outbound again ?", QMessageBox.Yes | QMessageBox.No)
                 if reply == QMessageBox.Yes:
                    break
                 else:
                    return
        outbound_orders = self.merge_outbound_orders(outbound_orders)
        print('outbound_orders: ', outbound_orders)
        # self.save_outbound_excel_for_express(outbound_orders)
        # self.save_outbound_excel_for_express0(outbound_orders)
        # self.save_outbound_excel_for_express1(outbound_orders)
        self.print_express(outbound_orders)


    def save_outbound_excel_for_express0(self, outbound_orders):
        alias = self.express_combo.currentText()
        config = get_express_config(alias)
        if not config:
            raise ValueError("no express config for alias: ", alias)
        express_name = config['name']
        default_shipper_zip = self.default_shipper_data['ShipperZIP']
        default_shipper_name = self.default_shipper_data['ShipperName']
        default_shipper_address = self.default_shipper_data['ShipperAddress']
        default_shipper_tel = self.default_shipper_data['ShipperTel']
        # 创建一个新的 Excel 工作簿，不显示 Excel 应用程序
        app = xw.App(visible=False)
        workbook = app.books.add()

        # 获取活动工作表
        worksheet = workbook.sheets.active

        # 定义所需的列标题
        # headers = [
        #     "出荷予定日", "送り状種類", "OrNO", "郵便番号", "住所1", "住所2", 
        #     # "住所3",
        #     "名前", "電話", "Comment1", "Comment2", "Text1", "口数",
        #     "発注番号", "Na", "YYU", "JYU", "JYU2", "TT"
        # ]
        headers = [
            "出荷予定日", "送り状種類", "OrNO", "郵便番号", "住所1", "住所2", 
            "名前", "電話", "Comment1", "Comment2", "Comment3", "Comment4", "口数",
            "SName", "SAddress1", "SAddress2", "STel", "SZIP", "D_Time", "D_Date"
        ]

        # 添加列标题到工作表
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cells(1, col_num)
            cell.value = header
            cell.api.HorizontalAlignment = -4108  # 居中对齐

            # 调整列宽以适应内容
            column_letter = chr(64 + col_num)
            worksheet.range(f"{column_letter}:{column_letter}").autofit()

        # 设置"出荷予定日"列的日期格式
        date_format = "yyyy/m/d"
        worksheet.range("A:A").number_format = date_format

        # 填充工作表中的订单数据
        for row_num, order in enumerate(outbound_orders, 2):
            # 获取当前日期
            current_date = datetime.date.today()
            formatted_date = current_date.strftime("%Y/%m/%d").replace("/0", "/")
            worksheet.cells(row_num, 1).value = formatted_date
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.cells(row_num, 2).value = "7"
            else:
                # sagawa
                worksheet.cells(row_num, 2).value = "0"
            worksheet.cells(row_num, 3).value = order.get("OrNO", "")
            worksheet.cells(row_num, 4).value = order.get("ZIP", "")
            if (express_name == Express.KURONEKOYAMATO.value):
                worksheet.cells(row_num, 5).value = order.get("Address", "")[0:16]
                worksheet.cells(row_num, 6).value = order.get("Address", "")[16:]
            else:
                worksheet.cells(row_num, 5).value = order.get("Address", "")
                worksheet.cells(row_num, 6).value = ""
            worksheet.cells(row_num, 7).value = order.get("Name", "")
            worksheet.cells(row_num, 8).value = order.get("TEL", "")
            worksheet.cells(row_num, 9).value = order.get("Comment1", "")
            worksheet.cells(row_num, 10).value = order.get("Comment2", "")
            worksheet.cells(row_num, 11).value = order.get("Comment3", "")
            worksheet.cells(row_num, 12).value = order.get("Comment4", "")
            worksheet.cells(row_num, 13).value = "1"
            worksheet.cells(row_num, 14).value = order.get("ShipperName", default_shipper_name)
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.cells(row_num, 15).value = order.get("ShipperAddress", default_shipper_address)[0:16]
                worksheet.cells(row_num, 16).value = order.get("ShipperAddress", default_shipper_address)[16:]
            else:
                worksheet.cells(row_num, 15).value = order.get("ShipperAddress", default_shipper_address)
                worksheet.cells(row_num, 16).value = ""
            worksheet.cells(row_num, 17).value = order.get("ShipperTel", default_shipper_tel)
            worksheet.cells(row_num, 18).value = order.get("ShipperZIP", default_shipper_zip)
            worksheet.cells(row_num, 19).value = order.get("D_Time", "")
            worksheet.cells(row_num, 20).value = order.get("D_Date", "")

        '''
        # 设置字体样式
        font = workbook.api.Font.Name = "ＭＳ Ｐゴシック"
        font_size = workbook.api.Font.Size = 12

        # 遍历所有单元格
        for row in worksheet.range(f"A1:{column_letter}{row_num}").api:
            for cell in row:
                cell.api.Font.Name = font
                cell.api.Font.Size = font_size
                cell.api.Value = unescape(cell.api.Value)
        '''

        # 保存工作簿
        workbook.save(ORDER_FOR_EXPRESS_PATH)
        workbook.close()

        # 关闭 Excel 应用程序
        app.quit()

    def save_outbound_excel_for_express(self, outbound_orders):
        alias = self.express_combo.currentText()
        config = get_express_config(alias)
        if not config:
            raise ValueError("no express config for alias: ", alias)
        express_name = config['name']
        default_shipper_zip = self.default_shipper_data['ShipperZIP']
        default_shipper_name = self.default_shipper_data['ShipperName']
        default_shipper_address = self.default_shipper_data['ShipperAddress']
        default_shipper_tel = self.default_shipper_data['ShipperTel']

        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        headers = [
            u"出荷予定日", u"送り状種類", u"OrNO", u"郵便番号", u"住所1", u"住所2",
            u"名前", u"電話", "Comment1", "Comment2", "Comment3", "Comment4", u"口数",
            "SName", "SAddress1", "SAddress2", "STel", "SZIP", "D_Time", "D_Date"
        ]

        # 添加列标题
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center')
            worksheet.column_dimensions[chr(col_num + 64)].auto_size = True

        # 设置日期格式
        date_format = "yyyy/m/d"
        worksheet.column_dimensions['A'].number_format = date_format

        # 填充订单数据
        for row_num, order in enumerate(outbound_orders, 2):
            current_date = datetime.date.today()
            formatted_date = current_date.strftime("%Y/%m/%d").replace("/0", "/")
            worksheet.cell(row=row_num, column=1, value=formatted_date)
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.cell(row=row_num, column=2, value="7")
            else:
                # sagawa
                worksheet.cell(row=row_num, column=2, value="0")
            worksheet.cell(row=row_num, column=3, value=order.get("OrNO", ""))
            worksheet.cell(row=row_num, column=4, value=order.get("ZIP", ""))
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.cell(row=row_num, column=5, value=order.get("Address", "")[:16])
                worksheet.cell(row=row_num, column=6, value=order.get("Address", "")[16:])
            else:
                worksheet.cell(row=row_num, column=5, value=order.get("Address", ""))
                worksheet.cell(row=row_num, column=6, value="")
            worksheet.cell(row=row_num, column=7, value=order.get("Name", ""))
            worksheet.cell(row=row_num, column=8, value=order.get("TEL", ""))
            worksheet.cell(row=row_num, column=9, value=order.get("Comment1", ""))
            worksheet.cell(row=row_num, column=10, value=order.get("Comment2", ""))
            worksheet.cell(row=row_num, column=11, value=order.get("Comment3", ""))
            worksheet.cell(row=row_num, column=12, value=order.get("Comment4", ""))
            worksheet.cell(row=row_num, column=13, value="1")
            worksheet.cell(row=row_num, column=14, value=order.get("ShipperName", default_shipper_name))
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.cell(row=row_num, column=15, value=order.get("ShipperAddress", default_shipper_address)[0:16])
                worksheet.cell(row=row_num, column=16, value=order.get("ShipperAddress", default_shipper_address)[16:])
            else:
                worksheet.cell(row=row_num, column=15, value=order.get("ShipperAddress", default_shipper_address))
                worksheet.cell(row=row_num, column=16, value="")
            worksheet.cell(row=row_num, column=17, value=order.get("ShipperTel", default_shipper_tel))
            worksheet.cell(row=row_num, column=18, value=order.get("ShipperZIP", default_shipper_zip))
            worksheet.cell(row=row_num, column=19, value=order.get("D_Time", ""))
            worksheet.cell(row=row_num, column=20, value=order.get("D_Date", ""))

        workbook.save(ORDER_FOR_EXPRESS_PATH)
        workbook.close()

    def save_outbound_excel_for_express1(self, outbound_orders):
        alias = self.express_combo.currentText()
        config = get_express_config(alias)
        if not config:
            raise ValueError("no express config for alias: ", alias)
        express_name = config['name']
        default_shipper_zip = self.default_shipper_data['ShipperZIP']
        default_shipper_name = self.default_shipper_data['ShipperName']
        default_shipper_address = self.default_shipper_data['ShipperAddress']
        default_shipper_tel = self.default_shipper_data['ShipperTel']

        # 创建一个新的工作簿
        workbook = xlsxwriter.Workbook(ORDER_FOR_EXPRESS_PATH)
        worksheet = workbook.add_worksheet()

        headers = [
            u"出荷予定日", u"送り状種類", u"OrNO", u"郵便番号", u"住所1", u"住所2",
            u"名前", u"電話", "Comment1", "Comment2", "Comment3", "Comment4", u"口数",
            "SName", "SAddress1", "SAddress2", "STel", "SZIP", "D_Time", "D_Date"
        ]

        # 添加列标题
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)
            worksheet.set_column(col_num, col_num, len(header))

        # 设置日期格式
        date_format = workbook.add_format({'num_format': 'yyyy/m/d'})
        worksheet.set_column('A:A', None, date_format)

        # 填充订单数据
        for row_num, order in enumerate(outbound_orders, 1):
            current_date = datetime.date.today()
            formatted_date = current_date.strftime("%Y/%m/%d").replace("/0", "/")
            worksheet.write(row_num, 0, formatted_date)
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.write(row_num, 1, "7")
            else:
                worksheet.write(row_num, 1, "0")
            worksheet.write(row_num, 2, order.get("OrNO", ""))
            worksheet.write(row_num, 3, order.get("ZIP", ""))
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.write(row_num, 4, order.get("Address", "")[:16])
                worksheet.write(row_num, 5, order.get("Address", "")[16:])
            else:
                worksheet.write(row_num, 4, order.get("Address", ""))
                worksheet.write(row_num, 5, "")
            worksheet.write(row_num, 6, order.get("Name", ""))
            worksheet.write(row_num, 7, order.get("TEL", ""))
            worksheet.write(row_num, 8, order.get("Comment1", ""))
            worksheet.write(row_num, 9, order.get("Comment2", ""))
            worksheet.write(row_num, 10, order.get("Comment3", ""))
            worksheet.write(row_num, 11, order.get("Comment4", ""))
            worksheet.write(row_num, 12, "1")
            worksheet.write(row_num, 13, order.get("ShipperName", default_shipper_name))
            if express_name == Express.KURONEKOYAMATO.value:
                worksheet.write(row_num, 14, order.get("ShipperAddress", default_shipper_address)[0:16])
                worksheet.write(row_num, 15, order.get("ShipperAddress", default_shipper_address)[16:])
            else:
                worksheet.write(row_num, 14, order.get("ShipperAddress", default_shipper_address))
                worksheet.write(row_num, 15, "")
            worksheet.write(row_num, 16, order.get("ShipperTel", default_shipper_tel))
            worksheet.write(row_num, 17, order.get("ShipperZIP", default_shipper_zip))
            worksheet.write(row_num, 18, order.get("D_Time", ""))
            worksheet.write(row_num, 19, order.get("D_Date", ""))

        workbook.close()

    def merge_outbound_orders(self, outbound_orders):
        merged_orders = {}
        for order in outbound_orders:
            picking_no = order['PickingNo']
            if picking_no not in merged_orders:
                merged_orders[picking_no] = order.copy()
                merged_orders[picking_no]['Comment2'] = ''
                merged_orders[picking_no]['Comment3'] = ''
            jan = order.get('JAN')
            if jan:
                comment2 = merged_orders[picking_no]['Comment2']
                comment3 = merged_orders[picking_no]['Comment3']
                jan_count_comment2 = len(comment2.split(' ')) if comment2 else 0
                jan_count_comment3 = len(comment3.split(' ')) if comment3 else 0
                if jan_count_comment2 <= jan_count_comment3:
                    merged_orders[picking_no]['Comment2'] = f"{comment2} {jan}"
                else:
                    merged_orders[picking_no]['Comment3'] = f"{comment3} {jan}"

        merged_orders_list = list(merged_orders.values())
        return merged_orders_list        


